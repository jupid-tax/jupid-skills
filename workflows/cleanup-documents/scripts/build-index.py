#!/usr/bin/env python3
"""Build a spreadsheet index of classified documents from manifest.json.

Usage:
    python3 build-index.py manifest.json --xlsx "Document Index.xlsx" --moves moves-result-<ts>.json
    python3 build-index.py manifest.json --csv document-index.csv --moves moves-result-<ts>.json --lang ru

Excel output requires `openpyxl` (pip install openpyxl); if it is missing the
script says so and writes CSV instead. CSV needs nothing. --moves points to
the moves-result-<timestamp>.json written by apply-move-plan.py --execute, so
the "Filed To" column shows the real final path (with any collision suffixes).
--lang ru localizes column headers and sheet names (default: en).
Existing files at the output path are never overwritten — the index gets a
" (2)" suffix instead.
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path

COLUMNS = [
    ("file", "Original File"),
    ("filed_to", "Filed To"),
    ("section", "Section"),
    ("subcategory", "Subcategory"),
    ("category", "Category (full path)"),
    ("doc_type", "Type"),
    ("counterparty", "Counterparty"),
    ("doc_date", "Date"),
    ("amount", "Amount"),
    ("currency", "Currency"),
    ("basis", "Basis for Classification"),
    ("needs_review", "Needs Review"),
    ("review_reason", "Review Reason"),
    ("duplicate_group", "Duplicate Group"),
    ("source", "Read From"),
]

HEADERS_RU = {
    "Original File": "Исходный файл",
    "Filed To": "Куда разложен",
    "Section": "Раздел",
    "Subcategory": "Подкатегория",
    "Category (full path)": "Категория (полный путь)",
    "Type": "Тип документа",
    "Counterparty": "Контрагент",
    "Date": "Дата документа",
    "Amount": "Сумма",
    "Currency": "Валюта",
    "Basis for Classification": "Основание классификации",
    "Needs Review": "Требует проверки",
    "Review Reason": "Причина проверки",
    "Duplicate Group": "Группа дубликатов",
    "Read From": "Способ чтения",
}

SHEETS = {
    "en": {"documents": "Documents", "summary": "Summary",
           "category": "Category", "files": "Files", "total": "Total",
           "needs_review": "Needs review", "yes": "YES"},
    "ru": {"documents": "Документы", "summary": "Сводка",
           "category": "Категория", "files": "Файлов", "total": "Всего",
           "needs_review": "Требует проверки", "yes": "ДА"},
}


def header_labels(lang: str) -> list:
    labels = [label for _, label in COLUMNS]
    if lang == "ru":
        return [HEADERS_RU[label] for label in labels]
    return labels


def no_overwrite(out: Path) -> Path:
    n = 1
    candidate = out
    while candidate.exists():
        n += 1
        candidate = out.with_name(f"{out.stem} ({n}){out.suffix}")
    return candidate


def load_rows(manifest_path: Path, moves_path: Path | None, lang: str) -> list:
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    final_paths = {}
    if moves_path:
        final_paths = {m["src"]: m["final"]
                       for m in json.loads(moves_path.read_text(encoding="utf-8"))}
    rows = []
    for entry in manifest:
        row = dict(entry)
        row["filed_to"] = final_paths.get(
            entry["file"], f"{entry['category']}/{Path(entry['file']).name}")
        row.setdefault("section", entry.get("category"))
        row.setdefault("subcategory", None)
        row["needs_review"] = SHEETS[lang]["yes"] if entry.get("needs_review") else ""
        rows.append(row)
    return rows


def summary_counts(rows: list) -> list:
    counts: dict = {}
    for row in rows:
        counts[row["category"]] = counts.get(row["category"], 0) + 1
    return sorted(counts.items())


def write_csv(rows: list, out: Path, lang: str) -> None:
    out = no_overwrite(out)
    with out.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(header_labels(lang))
        for row in rows:
            writer.writerow([row.get(key, "") if row.get(key) is not None else ""
                             for key, _ in COLUMNS])
    print(f"CSV index → {out} ({len(rows)} rows)")


def write_xlsx(rows: list, out: Path, lang: str) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False

    labels = SHEETS[lang]
    out = no_overwrite(out)
    wb = Workbook()
    ws = wb.active
    ws.title = labels["documents"]
    bold = Font(bold=True)

    ws.append(header_labels(lang))
    for cell in ws[1]:
        cell.font = bold
    for row in rows:
        ws.append([row.get(key) if row.get(key) is not None else ""
                   for key, _ in COLUMNS])
    widths = [40, 45, 24, 22, 30, 16, 24, 12, 12, 9, 60, 14, 30, 14, 12]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    summary = wb.create_sheet(labels["summary"])
    summary.append([labels["category"], labels["files"]])
    for cell in summary[1]:
        cell.font = bold
    for category, count in summary_counts(rows):
        summary.append([category, count])
    summary.append([])
    summary.append([labels["total"], len(rows)])
    summary.append([labels["needs_review"],
                    sum(1 for r in rows if r["needs_review"])])
    summary.column_dimensions["A"].width = 40

    wb.save(out)
    print(f"Excel index → {out} ({len(rows)} rows + {labels['summary']} sheet)")
    return True


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("manifest", help="manifest.json produced by classification")
    ap.add_argument("--xlsx", metavar="OUT", help="write Excel index")
    ap.add_argument("--csv", metavar="OUT", help="write CSV index")
    ap.add_argument("--moves", metavar="MOVES",
                    help="moves-result-<ts>.json from apply-move-plan.py --execute")
    ap.add_argument("--lang", choices=["en", "ru"], default="en",
                    help="language for headers and sheet names")
    args = ap.parse_args()

    if not args.xlsx and not args.csv:
        args.csv = "document-index.csv"

    rows = load_rows(Path(args.manifest),
                     Path(args.moves) if args.moves else None, args.lang)
    if not rows:
        print("manifest is empty — nothing to index", file=sys.stderr)
        return 1

    if args.xlsx:
        if not write_xlsx(rows, Path(args.xlsx), args.lang):
            fallback = Path(args.xlsx).with_suffix(".csv")
            print("openpyxl not installed (pip install openpyxl) — writing CSV instead")
            write_csv(rows, fallback, args.lang)
    if args.csv:
        write_csv(rows, Path(args.csv), args.lang)
    return 0


if __name__ == "__main__":
    sys.exit(main())
